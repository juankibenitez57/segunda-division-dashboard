"""
Exportación de datos a CSV y Excel con formato.

Excel generado:
  - Hoja "Todos"   : todos los movimientos
  - Hoja "YYYY-YY" : una hoja por cada temporada
Formato de todas las hojas:
  - Filtros automáticos
  - Primera fila congelada
  - Ancho de columnas ajustado al contenido
  - importe_numerico como número real de Excel (no texto)
"""
import logging
from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Border, Font, numbers, PatternFill, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

COLUMNS = [
    "temporada", "club", "jugador", "movimiento",
    "club_origen", "club_destino", "pais_club",
    "fecha", "importe_original", "importe_numerico", "tipo_operacion",
    "valor_mercado", "posicion", "edad", "nacionalidad",
]

# Columnas que deben mantenerse numéricas (no convertir a string)
NUMERIC_COLS = {"importe_numerico"}

DEDUP_KEYS = ["temporada", "club", "jugador", "movimiento"]


# ── DataFrame ─────────────────────────────────────────────────────────────────

def build_dataframe(records: list[dict]) -> pd.DataFrame:
    if not records:
        logger.warning("Sin registros para construir el DataFrame")
        return pd.DataFrame(columns=COLUMNS)

    df = pd.DataFrame(records)

    for col in COLUMNS:
        if col not in df.columns:
            df[col] = None if col in NUMERIC_COLS else "-"
    df = df[COLUMNS]

    # Eliminar duplicados
    before = len(df)
    df = df.drop_duplicates(subset=DEDUP_KEYS)
    if (removed := before - len(df)):
        logger.info(f"Eliminados {removed} registros duplicados")

    # Asegurar tipo numérico en importe_numerico (NaN donde no hay valor)
    df["importe_numerico"] = pd.to_numeric(df["importe_numerico"], errors="coerce")

    # Rellenar vacíos en columnas de texto
    text_cols = [c for c in COLUMNS if c not in NUMERIC_COLS]
    df[text_cols] = df[text_cols].fillna("-").replace("", "-")

    df = df.sort_values(["temporada", "club", "movimiento", "jugador"]).reset_index(drop=True)
    return df


# ── CSV ───────────────────────────────────────────────────────────────────────

def to_csv(df: pd.DataFrame, path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    df.to_csv(path, index=False, encoding="utf-8-sig")
    logger.info(f"CSV  → {path}  ({len(df):,} filas)")


# ── Excel ─────────────────────────────────────────────────────────────────────

def to_excel(df: pd.DataFrame, path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)

    seasons = sorted(df["temporada"].dropna().unique().tolist())

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        # Hoja 1: Todos los movimientos
        df.to_excel(writer, index=False, sheet_name="Todos")
        _format_sheet(writer.sheets["Todos"], df)

        # Una hoja por temporada
        for season in seasons:
            season_df = df[df["temporada"] == season].reset_index(drop=True)
            season_df.to_excel(writer, index=False, sheet_name=season)
            _format_sheet(writer.sheets[season], season_df)

    n_sheets = 1 + len(seasons)
    logger.info(f"Excel → {path}  ({len(df):,} filas, {n_sheets} hojas)")


# ── Formato ───────────────────────────────────────────────────────────────────

_HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
_EVEN_FILL   = PatternFill("solid", fgColor="DCE6F1")
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
_BODY_FONT   = Font(size=10)
_THIN_SIDE   = Side(style="thin", color="BFBFBF")
_BORDER      = Border(left=_THIN_SIDE, right=_THIN_SIDE, top=_THIN_SIDE, bottom=_THIN_SIDE)
_CENTER      = Alignment(horizontal="center", vertical="center", wrap_text=True)
_LEFT        = Alignment(horizontal="left", vertical="center")
_RIGHT       = Alignment(horizontal="right", vertical="center")

# Formato numérico: separador de miles, sin decimales
_NUM_FORMAT  = "#,##0"


def _format_sheet(ws, df: pd.DataFrame) -> None:
    n_rows   = len(df)
    n_cols   = len(df.columns)
    col_names = list(df.columns)

    # Índice (1-based) de la columna numérica
    num_col_idx = (col_names.index("importe_numerico") + 1
                   if "importe_numerico" in col_names else None)

    # ── Cabecera ──────────────────────────────────────────────────────────────
    for col_idx in range(1, n_cols + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font      = _HEADER_FONT
        cell.fill      = _HEADER_FILL
        cell.alignment = _CENTER
        cell.border    = _BORDER

    # ── Filas de datos ────────────────────────────────────────────────────────
    for row_idx in range(2, n_rows + 2):
        for col_idx in range(1, n_cols + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font   = _BODY_FONT
            cell.border = _BORDER
            if row_idx % 2 == 0:
                cell.fill = _EVEN_FILL

            if col_idx == num_col_idx:
                # Número real + alineación derecha + formato con separador de miles
                cell.alignment    = _RIGHT
                cell.number_format = _NUM_FORMAT
            else:
                cell.alignment = _LEFT

    # ── Ancho de columnas (ajuste automático al contenido) ────────────────────
    for col_idx, col_name in enumerate(col_names, 1):
        letter = get_column_letter(col_idx)

        # Longitud máxima entre cabecera y datos
        header_len = len(col_name)
        if col_name in df.columns and len(df) > 0:
            # Para numéricos, usar la representación formateada con separadores
            if col_name in NUMERIC_COLS:
                max_data_len = (
                    df[col_name]
                    .dropna()
                    .map(lambda x: len(f"{x:,.0f}"))
                    .max() if df[col_name].notna().any() else 0
                )
            else:
                max_data_len = df[col_name].astype(str).map(len).max()
        else:
            max_data_len = 0

        width = min(max(int(max_data_len), header_len) + 3, 45)
        ws.column_dimensions[letter].width = width

    # ── Fila fija + filtro automático ─────────────────────────────────────────
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(n_cols)}1"
